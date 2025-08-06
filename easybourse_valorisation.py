import os
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import logging
import re

# Configuration du logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EasyBourseValorisationDownloader:
    def __init__(self, username, password, download_dir=None):
        self.username = username
        self.password = password
        self.base_url = "https://www.easybourse.com"
        self.valorisation_url = f"{self.base_url}/secure/compte/valorisation"

        # Définir le répertoire de téléchargement
        if download_dir is None:
            self.download_dir = os.path.dirname(os.path.abspath(__file__))
        else:
            self.download_dir = download_dir

        logger.info(f"Répertoire de téléchargement: {self.download_dir}")

    def setup_driver(self):
        """Configure et retourne un driver Selenium avec téléchargement automatique"""
        try:
            options = Options()

            # Configuration pour le téléchargement automatique
            prefs = {
                "download.default_directory": self.download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            }
            options.add_experimental_option("prefs", prefs)

            # Options pour éviter la détection
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            # Autres options utiles
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            # options.add_argument('--headless')  # Décommenter pour mode headless

            driver = webdriver.Chrome(options=options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            driver.set_window_size(1920, 1080)

            return driver
        except Exception as e:
            logger.error(f"Erreur lors de la configuration du driver: {e}")
            raise

    def login(self, driver):
        """Se connecte à EasyBourse"""
        try:
            # Étape 1: Page de connexion - Entrer l'identifiant
            logger.info("Navigation vers la page de connexion...")
            driver.get(f"{self.base_url}/login")

            # Attendre que le champ username soit visible
            wait = WebDriverWait(driver, 10)
            username_field = wait.until(
                EC.presence_of_element_located((By.NAME, "username"))
            )

            # Accepter les cookies si présent
            try:
                logger.info("Acceptation des cookies...")
                time.sleep(2)
                driver.find_element(By.XPATH, "//button[contains(text(), 'Ok pour moi')]").click()
            except:
                pass

            logger.info("Saisie de l'identifiant...")
            username_field.send_keys(self.username)

            # Cliquer sur le bouton Continuer
            continue_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Continuer')]")
            continue_button.click()

            # Étape 2: Page du mot de passe
            logger.info("En attente de la page de mot de passe...")
            time.sleep(2)

            # Gérer le clavier virtuel ou le champ normal
            try:
                # Détecter le clavier virtuel
                virtual_keyboard = None
                for key_id in range(1000):
                    try:
                        elements = driver.find_elements(By.CLASS_NAME, f"jss{key_id}")
                        digits = [el.text.strip() for el in elements if
                                  el.text.strip().isdigit() and len(el.text.strip()) == 1]
                        if set(digits) == set("0123456789"):
                            virtual_keyboard = elements
                            logger.info(f"✅ Clavier virtuel détecté avec jss{key_id}")
                            break
                    except:
                        continue

                if virtual_keyboard:
                    logger.info("Utilisation du clavier virtuel...")
                    for digit in self.password:
                        for button in virtual_keyboard:
                            if button.text == digit:
                                button.click()
                                time.sleep(0.2)
                                break
                else:
                    # Essayer le champ de mot de passe normal
                    try:
                        password_field = driver.find_element(By.NAME, "password")
                        password_field.send_keys(self.password)
                    except:
                        logger.info("Veuillez saisir le mot de passe manuellement...")
                        time.sleep(30)

            except Exception as e:
                logger.error(f"Erreur lors de la saisie du mot de passe: {e}")

            # Cliquer sur Se connecter
            login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Se connecter')]")
            login_button.click()

            logger.info("Connexion en cours...")
            time.sleep(5)

            return True

        except Exception as e:
            logger.error(f"Erreur lors de la connexion: {e}")
            return False

    def download_valorisation_csv(self, driver):
        """Télécharge le fichier CSV de valorisation"""

        # URL de la page de téléchargement du CSV
        driver.get('https://www.easybourse.com/easybourse/secure/exportCsvValorisationTempsReel.html?siteLanguage=fr')
        files_before = set(os.listdir(self.download_dir))

        # Attendre que le fichier soit téléchargé
        timeout = 30
        start_time = time.time()

        while time.time() - start_time < timeout:
            files_after = set(os.listdir(self.download_dir))
            new_files = files_after - files_before

            # Chercher un fichier CSV nouvellement créé
            csv_files = [f for f in new_files if f.endswith('.csv')]

            if csv_files:
                csv_filename = csv_files[0]
                logger.info(f"Fichier CSV téléchargé: {csv_filename}")
                return os.path.join(self.download_dir, csv_filename)

            time.sleep(1)

        return None

    def parse_csv_data(self, csv_path):
        """Parse le fichier CSV et extrait les données avec les totaux comme colonnes"""
        try:
            # Lire le fichier avec l'encodage approprié
            with open(csv_path, 'r', encoding='cp1252') as f:
                content = f.read()

            lines = content.split('\n')

            # Extraire la date de valorisation (ligne 2)
            date_match = re.search(r'Valorisation au;(\d{2}/\d{2}/\d{4})', lines[2] if len(lines) > 2 else '')
            if date_match:
                date_str = date_match.group(1)
                valorisation_date = datetime.strptime(date_str, '%d/%m/%Y')
                logger.info(f"Date de valorisation: {valorisation_date.strftime('%d/%m/%Y')}")
            else:
                valorisation_date = datetime.now()
                logger.warning("Date de valorisation non trouvée, utilisation de la date actuelle")

            # Extraire les totaux du CSV (lignes 7-12)
            totals_dict = {}

            logger.info("Extraction des totaux...")
            for i in range(7, 13):  # Lignes 7 à 12
                if i < len(lines) and lines[i].strip():
                    parts = lines[i].split(';')
                    if len(parts) >= 2 and parts[0].strip():
                        label = parts[0].strip()
                        montant = parts[1].strip() if len(parts) > 1 else '0'

                        try:
                            # Convertir le montant
                            montant_float = float(montant.replace(',', '.').replace(' ', ''))

                            # Mapper les labels aux noms de colonnes
                            if label == 'Total positions sous dossier':
                                totals_dict['Total positions sous dossier'] = montant_float
                            elif label == 'Solde espèces':
                                totals_dict['Solde espèces'] = montant_float
                            elif label == 'Valeur totale':
                                totals_dict['Valeur totale'] = montant_float

                            logger.info(f"  • {label}: {montant_float:,.2f}€")

                        except ValueError as e:
                            logger.warning(f"Impossible de convertir la valeur pour {label}: {montant}")

            # Trouver le début du tableau des positions
            header_index = -1
            for i, line in enumerate(lines):
                if 'Valeur;Code Isin;Place de cotation' in line:
                    header_index = i
                    break

            if header_index == -1:
                logger.error("En-têtes du tableau non trouvés")
                return None

            # Extraire les positions
            positions_lines = []
            for i in range(header_index, len(lines)):
                if lines[i].strip() and ';' in lines[i]:
                    positions_lines.append(lines[i])

            # Parser les positions
            import io
            positions_content = '\n'.join(positions_lines)
            df = pd.read_csv(io.StringIO(positions_content), sep=';', decimal=',')

            # Nettoyer les colonnes
            df.columns = df.columns.str.strip()

            # Ajouter la date
            df['Date'] = valorisation_date

            # Convertir les colonnes numériques
            numeric_columns = ['Quantité', 'Cours', 'Prix moyen', 'Valorisation', '+/- value', 'Performance (%)',
                               'Poids']

            for col in numeric_columns:
                if col in df.columns:
                    def safe_convert(x):
                        if pd.isna(x):
                            return x
                        try:
                            x_str = str(x).strip()
                            x_str = x_str.replace(' ', '').replace(',', '.')
                            if col == 'Performance (%)' and '%' in x_str:
                                x_str = x_str.replace('%', '')
                            return float(x_str)
                        except:
                            return None

                    df[col] = df[col].apply(safe_convert)

            # IMPORTANT : Ajouter les totaux comme colonnes (même valeur pour toutes les lignes)
            for col_name, value in totals_dict.items():
                df[col_name] = value

            logger.info(f"Colonnes de totaux ajoutées: {list(totals_dict.keys())}")

            # Afficher un résumé
            logger.info(f"Données extraites: {len(df)} positions")
            if len(df) > 0:
                valorisation_totale = df['Valorisation'].sum()
                logger.info(f"Valorisation calculée des positions: {valorisation_totale:,.2f}€")
                if 'Total positions sous dossier' in totals_dict:
                    logger.info(f"Total positions (du CSV): {totals_dict['Total positions sous dossier']:,.2f}€")

            return df

        except Exception as e:
            logger.error(f"Erreur lors du parsing du CSV: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def update_excel(self, df_new, excel_path='EasyBourse.xlsx'):
        """Met à jour le fichier Excel avec les nouvelles données incluant les colonnes de totaux"""
        try:
            logger.info(f"📁 Mise à jour du fichier : {excel_path}")

            # Colonnes de totaux à gérer
            TOTAL_COLUMNS = ['Valeur totale', 'Total positions sous dossier', 'Solde espèces']

            # S'assurer que les colonnes de totaux existent dans df_new
            for col in TOTAL_COLUMNS:
                if col not in df_new.columns:
                    df_new[col] = None

            # Vérifier si le fichier Excel existe
            if os.path.exists(excel_path):
                # Lire le fichier existant
                df_existing = pd.read_excel(excel_path, sheet_name='Data')
                logger.info(f"📊 Fichier existant chargé : {len(df_existing)} lignes")

                # S'assurer que les colonnes de totaux existent dans df_existing
                for col in TOTAL_COLUMNS:
                    if col not in df_existing.columns:
                        df_existing[col] = None

                # S'assurer que les dates sont au bon format
                df_existing['Date'] = pd.to_datetime(df_existing['Date'])
                df_new['Date'] = pd.to_datetime(df_new['Date'])

                # Obtenir les dates uniques du nouveau CSV
                new_dates = df_new['Date'].unique()

                # Traiter chaque date du nouveau CSV
                for date in new_dates:
                    # Extraire les données pour cette date
                    df_date_new = df_new[df_new['Date'] == date].copy()

                    # Récupérer les valeurs des totaux pour cette date (même pour toutes les lignes)
                    totals_values = {}
                    for col in TOTAL_COLUMNS:
                        if col in df_date_new.columns and not df_date_new[col].isna().all():
                            totals_values[col] = df_date_new[col].iloc[0]  # Prendre la première valeur

                    logger.info(f"📅 Traitement de la date {date.strftime('%d/%m/%Y')}")
                    logger.info(f"   Totaux: {totals_values}")

                    # Vérifier si cette date existe déjà dans le fichier Excel
                    if date in df_existing['Date'].values:
                        logger.info(f"✅ Date {date.strftime('%d/%m/%Y')} existe déjà dans Excel")

                        # Pour chaque position de cette date dans le nouveau CSV
                        for idx, row in df_date_new.iterrows():
                            valeur = row['Valeur']

                            # Vérifier si cette action existe déjà pour cette date
                            mask = (df_existing['Date'] == date) & (df_existing['Valeur'] == valeur)

                            if mask.any():
                                # Mise à jour de la ligne existante
                                logger.info(f"   🔄 Mise à jour de {valeur}")
                                update_idx = df_existing[mask].index[0]

                                # Mettre à jour toutes les colonnes
                                for col in df_new.columns:
                                    df_existing.at[update_idx, col] = row[col]
                            else:
                                # Ajouter la nouvelle position
                                logger.info(f"   ➕ Ajout de {valeur}")

                                # Trouver où insérer la nouvelle ligne
                                date_mask = df_existing['Date'] == date
                                if date_mask.any():
                                    last_idx_for_date = df_existing[date_mask].index[-1]
                                    insert_idx = last_idx_for_date + 1
                                else:
                                    insert_idx = len(df_existing)

                                # Insérer la nouvelle ligne
                                df_existing = pd.concat([
                                    df_existing.iloc[:insert_idx],
                                    pd.DataFrame([row]),
                                    df_existing.iloc[insert_idx:]
                                ], ignore_index=True)

                        # Mettre à jour les colonnes de totaux pour toutes les lignes de cette date
                        date_mask = df_existing['Date'] == date
                        for col, value in totals_values.items():
                            df_existing.loc[date_mask, col] = value
                            logger.info(f"   📊 Mise à jour {col} = {value:,.2f} pour toutes les lignes")

                    else:
                        # Date n'existe pas, ajouter toutes les lignes
                        logger.info(f"🆕 Nouvelle date {date.strftime('%d/%m/%Y')}, ajout de {len(df_date_new)} lignes")

                        # Trouver où insérer selon l'ordre chronologique
                        if len(df_existing) > 0:
                            dates_existing = df_existing['Date'].unique()
                            dates_existing_sorted = sorted(dates_existing)

                            insert_position = None
                            for i, existing_date in enumerate(dates_existing_sorted):
                                if date < existing_date:
                                    first_idx_of_date = df_existing[df_existing['Date'] == existing_date].index[0]
                                    insert_position = first_idx_of_date
                                    break

                            if insert_position is None:
                                insert_position = len(df_existing)

                            # Insérer les nouvelles données
                            df_existing = pd.concat([
                                df_existing.iloc[:insert_position],
                                df_date_new,
                                df_existing.iloc[insert_position:]
                            ], ignore_index=True)
                        else:
                            df_existing = df_date_new

                df_combined = df_existing

            else:
                # Si le fichier n'existe pas, utiliser les nouvelles données
                df_combined = df_new
                logger.info(f"🆕 Création d'un nouveau fichier Excel avec {len(df_new)} lignes")

            # Trier par date puis par valeur
            df_combined = df_combined.sort_values(['Date', 'Valeur'], ascending=[True, True])

            # Réinitialiser l'index
            df_combined = df_combined.reset_index(drop=True)

            # Nettoyer df_combined : supprimer les colonnes Unnamed
            df_combined = df_combined.loc[:, ~df_combined.columns.str.contains('Unnamed')]

            logger.info(f"📊 Total après mise à jour: {len(df_combined)} lignes")

            # Sauvegarder dans Excel
            logger.info(f"💾 Sauvegarde dans : {excel_path}")
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_combined.to_excel(writer, sheet_name='Data', index=False)

                # Ajuster la largeur des colonnes
                worksheet = writer.sheets['Data']
                for idx, column in enumerate(df_combined.columns):
                    column_length = max(
                        df_combined[column].astype(str).map(len).max(),
                        len(column)
                    )
                    # Gérer les colonnes au-delà de Z
                    if idx < 26:
                        col_letter = chr(65 + idx)
                    else:
                        col_letter = chr(65 + idx // 26 - 1) + chr(65 + idx % 26)
                    worksheet.column_dimensions[col_letter].width = min(column_length + 2, 50)

                # Formater les colonnes de totaux avec une couleur de fond
                from openpyxl.styles import PatternFill, Font
                light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                bold_font = Font(bold=True)

                # Trouver les indices des colonnes de totaux
                for col_name in TOTAL_COLUMNS:
                    if col_name in df_combined.columns:
                        col_idx = df_combined.columns.get_loc(col_name) + 1  # +1 car Excel commence à 1

                        # Formater l'en-tête
                        header_cell = worksheet.cell(row=1, column=col_idx)
                        header_cell.fill = light_blue_fill
                        header_cell.font = bold_font

                        # Formater les données avec un fond léger
                        for row in range(2, len(df_combined) + 2):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.fill = light_blue_fill

            logger.info(f"✅ Fichier Excel mis à jour: {excel_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Erreur lors de la mise à jour du fichier Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def backup_excel(self, excel_path):
        """Crée une sauvegarde du fichier Excel avant modification"""
        if os.path.exists(excel_path):
            backup_dir = 'Save'
            os.makedirs(backup_dir, exist_ok=True)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(backup_dir, f'EasyBourse_backup_{timestamp}.xlsx')

            import shutil
            shutil.copy2(excel_path, backup_path)
            logger.info(f"Sauvegarde créée: {backup_path}")

            # Nettoyer les anciennes sauvegardes (garder seulement les 10 dernières)
            backups = sorted([f for f in os.listdir(backup_dir) if f.endswith('.xlsx')])
            if len(backups) > 10:
                for old_backup in backups[:-10]:
                    os.remove(os.path.join(backup_dir, old_backup))
                    logger.info(f"Ancienne sauvegarde supprimée: {old_backup}")

    def run(self, excel_path=None):
        """Exécute le processus complet"""
        driver = None
        try:
            # Définir le chemin du fichier Excel
            if excel_path is None:
                excel_path = 'EasyBourse.xlsx'

            # Créer le chemin absolu si nécessaire
            excel_path = os.path.abspath(excel_path)
            logger.info(f"Fichier Excel cible : {excel_path}")

            # Configurer le driver
            driver = self.setup_driver()

            # Se connecter
            if not self.login(driver):
                logger.error("Échec de la connexion")
                return False

            # Télécharger le CSV
            csv_path = self.download_valorisation_csv(driver)
            if not csv_path:
                logger.error("Échec du téléchargement du CSV")
                return False

            # Parser les données
            df = self.parse_csv_data(csv_path)
            if df is None:
                logger.error("Échec du parsing du CSV")
                return False

            # Créer une sauvegarde si le fichier existe
            if os.path.exists(excel_path):
                self.backup_excel(excel_path)

            # Mettre à jour Excel - IMPORTANT: Passer excel_path en paramètre !
            if self.update_excel(df, excel_path):  # <-- Correction ici
                logger.info("✅ Processus terminé avec succès!")

                # Optionnel : Supprimer le CSV téléchargé
                try:
                    os.remove(csv_path)
                    logger.info(f"Fichier CSV temporaire supprimé : {csv_path}")
                except Exception as e:
                    logger.warning(f"Impossible de supprimer le CSV : {e}")

                return True
            else:
                return False

        except Exception as e:
            logger.error(f"Erreur générale: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if driver:
                driver.quit()


# Utilisation du script
if __name__ == "__main__":
    # Remplacer par vos identifiants
    from logins import id, password

    USERNAME = id
    PASSWORD = password

    # Créer et lancer le downloader
    downloader = EasyBourseValorisationDownloader(USERNAME, PASSWORD)
    downloader.run()